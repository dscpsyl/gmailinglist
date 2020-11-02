import pickle
import base64
import mimetypes
import time
import sys
import os
import datetime as dt
import openpyxl as xl
import json as js
import logging as lg 
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from apiclient import errors
from datetime import datetime as dt

#* Turns on and sets Parameters
debug = True
safeMode = True
recordSent = False
showBody = True
limit = 100

#* If modifying these scopes, delete the file token.pickle in login()
SCOPES = ['https://www.googleapis.com/auth/gmail.compose']

#* Returns service 
def login():
  creds = None
  if os.path.exists('token.pickle'):
    lg.info("Token found.")
    with open('token.pickle', 'rb') as token:
        creds = pickle.load(token)
        lg.debug("Toekn inport successful")
  if not creds or not creds.valid:
    if creds and creds.expired and creds.refresh_token:
      lg.debug("Token has expired. Refreshing request...")
      creds.refresh(Request())
      lg.debug("Token refresh successful.")
    else:
        lg.info("No valid tokens found.")
        lg.debug("Logging into account...")
        flow = InstalledAppFlow.from_client_secrets_file(
            'credentials.json', SCOPES)
        creds = flow.run_local_server(port=0)
        lg.info("Login successful.")
    with open('token.pickle', 'wb') as token:
      lg.debug("Saving token...")
      pickle.dump(creds, token)
          
  service = build('gmail', 'v1', credentials=creds)
  lg.debug("Service build successful!")
  return service 

#* Returns the "raw" value of the draft object
def create_message(sender, to, subject, message_text):
  message = MIMEText(message_text)
  message['to'] = to
  message['from'] = sender
  message['subject'] = subject
  message = base64.urlsafe_b64encode(message.as_bytes())
  message = message.decode()
  lg.debug("     Message created...")
  return {'raw': message }

#* Returns the "raw" value of the draft object
def create_message_with_attachment(sender, to, subject, message_text, file):
  message = MIMEMultipart()
  message['to'] = to
  message['from'] = sender
  message['subject'] = subject

  msg = MIMEText(message_text)
  message.attach(msg)

  content_type, encoding = mimetypes.guess_type(file)

  if content_type is None or encoding is not None:
    content_type = 'application/octet-stream'
  main_type, sub_type = content_type.split('/', 1)
  if main_type == 'text':
    fp = open(file, 'rb')
    msg = MIMEText(fp.read(), _subtype=sub_type)
    fp.close()
  elif main_type == 'image':
    fp = open(file, 'rb')
    msg = MIMEImage(fp.read(), _subtype=sub_type)
    fp.close()
  elif main_type == 'audio':
    fp = open(file, 'rb')
    msg = MIMEAudio(fp.read(), _subtype=sub_type)
    fp.close()
  else:
    fp = open(file, 'rb')
    msg = MIMEBase(main_type, sub_type)
    msg.set_payload(fp.read())
    fp.close()
  filename = os.path.basename(file)
  msg.add_header('Content-Disposition', 'attachment', filename=filename)
  message.attach(msg)

  message = base64.urlsafe_b64encode(message.as_bytes())
  message = message.decode()
  lg.debug("     Message with attachment " + file + " created...")
  return {'raw': message }

#* Connects to API and creates draft. Returns the draftID
def create_draft(service, user_id, message_body):
  try:
    message = {'message': message_body}
    
    lg.debug("Sending draft to Google's servers...")
    draft = service.users().drafts().create(userId=user_id, body=message).execute()

    lg.debug('Draft id: %s\nDraft message: %s' % (draft['id'], draft['message']))
    lg.info("     Draft saved to Google's servers with id: " + draft["id"])

    return draft["id"]
  except errors.HttpError as error:
    lg.critical('An error occurred and draft could not be saved: %s' % error)
    if __name__ == "__main__":
      sys.exit(1)
    else:
      return None

#* Connects to API and sends draft by ID. Returns TRUE or FALSE depending on success
def send_draft(service, user_Id, draftId, to, name):
  areYouSure = "Y"
  if safeMode == True:
    print("You are sending to: %s with name %s" % (to, name))
    areYouSure = input("Are you sure? (Y/n): ")
  
  if areYouSure != "Y":
    lg.error("User has stopped this message from sending. The draft is still saved in the account and can still be accessed anytime.")
  else:
    try:
        lg.debug("Sending draft to recipient...")
        service.users().drafts().send(userId=user_Id, body={ 'id': draftId }).execute()

        lg.info('     Draft id: %s has been sent to %s successfully!' % (draftId, to))
        return True
    except errors.HttpError as error:
        lg.error('An error occurred and draft could not be sent: %s' % error)
        if __name__ == "__main__":
          sys.exit(1)
        else:
          return False





##* CHANGE THIS PART FOR DIFFERNET USAGES DEPENDING ON YOUR .XSLX FILE
def process(ws, service, wb, wbname):
  userId="me" #TODO
  jsFile = "data.json"
  try:
    with open(jsFile):
      jsonData = js.load(jsFile)
  except Exception as e:
    lg.critical("Json File Error. Please check log")
    if debug == True:
      print("Json File Error. Please check log")
    sys.exit(1)
    
  lg.debug("Using userId: " + userId)
  sender =  jsonData["name"] + "<" + jsonData["sender"] + ">" 
  lg.info("Using sender as: " + sender)
  maxRow = ws.max_row
  maxCol = ws.max_column
  lg.info("Max rows is: " + str(maxRow) + "and max column is: " + str(maxCol))
  
  for rows in range(maxRow):
    if rows == 0 or rows == 1:
      continue
    
    
    
    subject = jsonData["subject"] 
    lg.debug("Subject is: " + subject) 
    recieptName = ws["A" + str(rows)].value
    lg.debug("Recipient is: " + recieptName)
    to = ws["B" + str(rows)].value
    lg.debug("Sending to: " + to)
    body = 
    if showBody == True:
      lg.info("Body content is: " + body)
    
    
    
    
    
    email(to, recieptName, body, service, userId, sender, subject)
    lg.debug("Sending email to: " + recieptName + " with the address: " + to + " and body containing:\n")
    if debug == True:
      print("Sending email to: " + recieptName + " with the address: " + to + " and body containing:\n")
    time.sleep(1)
    if recordSent == True:
      greenfill = xl.styles.PatternFill(start_color='00FF00', end_color='000000', fill_type='solid')
      recordSentMail(rows, ws, greenfill, maxCol)
    if rows == limit:
      lg.critical("Max email limit reached. Exiting...")
      if recordSent == True:
        wb.save(wbname + ".new")
        os.remove(wbname)
        os.rename(wbname + ".new", wbname)
      if debug == True:
        print("Max email limit reached. Exiting...")
      sys.exit(0)
    rows += 1
    
  lg.info("List has been cycled through. Exiting...")
  if recordSent == True:
    wb.save(wbname + ".new")
    os.remove(wbname)
    os.rename(wbname + ".new", wbname)
  if debug == True:
    print("List has been cycled through. Exiting...")
 
#* Changes sent address rows to green and puts a date/time stamp next to it
def recordSentMail(row, ws, color, maxCol):
  maxColMax = maxCol + 1
  colLttrMax = xl.utils.get_column_letter(maxColMax)
  now = dt.now()
  now_string = now.strftime("%d/%m/%Y-%H:%M:%S")
  for col in range(maxCol+1):
    if col == 0:
      continue
    colLttr = xl.utils.get_column_letter(col)
    ws[colLttr + str(row)].fill = color
  ws[colLttrMax + str(row)].value = now_string
  lg.debug("Time for row: " + str(row) + " is assigned to: " + str(now_string))

#* Starts the email API process
def email(to, recieptName, body, service, userId, sender, subject):
  
  lg.info("Starting on email for: " + recieptName)
  message = create_message(sender, to, subject, body)
  draftId = create_draft(service, userId, message)
  send_draft(service, userId, draftId, to, recieptName)
  lg.info("Email for %s has been successfully sent to %s!" % (recieptName, to))
  lg.info("#####")

#* Starts Logging
def scriptLog():
  #*check for old lenghty log | Creates log folder and file
  date = dt.now().date()
  currentDir = os.getcwd()
  logs = currentDir + "/logs"
  
  if debug == True:
    loggingLvl = lg.DEBUG 
  else:
    loggingLvl = lg.INFO
    
  try:
    os.chdir(logs)
  except FileNotFoundError:
    if debug == True:
      print("No Log folder found. Creating one now...")
    os.mkdir(logs)
    lg.basicConfig(level=loggingLvl, filename="logs/massEmail" + str(date) + ".log", format="%(asctime)s:%(levelname)s:%(message)s", datefmt='%d-%b-%y %H:%M:%S')
      
  os.chdir(currentDir)
  lg.basicConfig(level=loggingLvl, filename="logs/massEmail" + str(date) + ".log", format="%(asctime)s:%(levelname)s:%(message)s", datefmt="%d-%b-%y %H:%M:%S")

  lg.info("********************************************")
  lg.info("Starting Gmail email script...")
  if safeMode == True:
    lg.warning("SAFEMODE is ON")
      
#* Main process...come on you can figure this out.
def main(workbook):
  scriptLog()
  service = login()

  try:
    wbname = workbook
    wb = xl.load_workbook(filename=wbname)
    lg.debug("Workoutbook found!")
  except:
    lg.critical("The workbook cannot be found.")
    if debug == True:
      print("Workbook cannot be found.")
    sys.exit(1)
      
  ws = wb.active
  lg.debug("Sheet found")
  process(ws, service, wb, wbname)
#################################################
if __name__ == '__main__':
  wbname = input("Where is the input list?: ")
  main(wbname)